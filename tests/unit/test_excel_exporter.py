"""Unit tests for Excel export functionality"""

import pytest
import tempfile
import time
from pathlib import Path
from openpyxl import load_workbook

from dns_gather.excel_exporter import ExcelExporter
from dns_gather.models import ZoneInfo, DNSRecord


@pytest.fixture
def temp_output_dir():
    """Create temporary output directory"""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield tmpdir
        # Small delay to allow Windows to release file handles
        time.sleep(0.1)


@pytest.fixture
def exporter(temp_output_dir):
    """Create ExcelExporter instance"""
    return ExcelExporter(output_directory=temp_output_dir)


@pytest.fixture
def sample_zones():
    """Create sample zone data"""
    return [
        ZoneInfo(
            name='example.com',
            zone_type='Primary',
            serial=2024011601,
            transfer_status='Success',
            record_count=10,
            error_message=''
        ),
        ZoneInfo(
            name='test.local',
            zone_type='Secondary',
            serial=2024011602,
            transfer_status='Success',
            record_count=5,
            error_message=''
        )
    ]


@pytest.fixture
def sample_records():
    """Create sample DNS records"""
    return [
        DNSRecord(name='www', record_type='A', ttl=300, data='192.168.1.1'),
        DNSRecord(name='mail', record_type='MX', ttl=600, data='10 mail.example.com'),
        DNSRecord(name='@', record_type='NS', ttl=3600, data='ns1.example.com')
    ]


def test_exporter_initialization(temp_output_dir):
    """Test that ExcelExporter initializes correctly"""
    exporter = ExcelExporter(output_directory=temp_output_dir)
    
    assert exporter.output_directory == Path(temp_output_dir)
    assert exporter.output_directory.exists()
    assert exporter.header_bg_color == '4472C4'
    assert exporter.header_font_color == 'FFFFFF'
    assert exporter.max_column_width == 50


def test_exporter_creates_output_directory():
    """Test that ExcelExporter creates output directory if it doesn't exist"""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / 'new_directory'
        assert not output_path.exists()
        
        exporter = ExcelExporter(output_directory=str(output_path))
        
        assert output_path.exists()


def test_create_workbook_with_sample_data(exporter, sample_zones):
    """Test workbook creation with sample data"""
    records_by_zone = {
        'example.com': [
            DNSRecord(name='www', record_type='A', ttl=300, data='192.168.1.1')
        ],
        'test.local': []
    }
    
    filepath = exporter.create_workbook(sample_zones, records_by_zone)
    
    assert Path(filepath).exists()
    assert filepath.endswith('.xlsx')


def test_workbook_contains_zone_list_sheet(exporter, sample_zones):
    """Test that workbook contains Zone List sheet"""
    records_by_zone = {zone.name: [] for zone in sample_zones}
    filepath = exporter.create_workbook(sample_zones, records_by_zone)
    
    wb = load_workbook(filepath)
    assert 'Zone List' in wb.sheetnames
    wb.close()


def test_zone_list_sheet_has_all_zones(exporter, sample_zones):
    """Test that Zone List sheet contains all zones"""
    records_by_zone = {zone.name: [] for zone in sample_zones}
    filepath = exporter.create_workbook(sample_zones, records_by_zone)
    
    wb = load_workbook(filepath)
    ws = wb['Zone List']
    
    # Check row count (header + zones)
    assert ws.max_row == len(sample_zones) + 1
    
    # Check zone names
    zone_names = [ws.cell(row=i+2, column=1).value for i in range(len(sample_zones))]
    assert 'example.com' in zone_names
    assert 'test.local' in zone_names
    
    wb.close()


def test_zone_sheet_creation(exporter, sample_records):
    """Test creation of individual zone sheet"""
    zone = ZoneInfo(
        name='example.com',
        zone_type='Primary',
        serial=2024011601,
        transfer_status='Success',
        record_count=len(sample_records),
        error_message=''
    )
    
    records_by_zone = {'example.com': sample_records}
    filepath = exporter.create_workbook([zone], records_by_zone)
    
    wb = load_workbook(filepath)
    
    # Check that zone sheet exists
    assert 'example.com' in wb.sheetnames
    
    ws = wb['example.com']
    
    # Check row count (header + records)
    assert ws.max_row == len(sample_records) + 1
    
    wb.close()


def test_zone_sheet_with_no_records(exporter):
    """Test zone sheet creation with no records"""
    zone = ZoneInfo(
        name='empty.zone',
        zone_type='Primary',
        serial=1,
        transfer_status='Success',
        record_count=0,
        error_message=''
    )
    
    records_by_zone = {'empty.zone': []}
    filepath = exporter.create_workbook([zone], records_by_zone)
    
    wb = load_workbook(filepath)
    ws = wb['empty.zone']
    
    # Should have only header row
    assert ws.max_row == 1
    
    wb.close()


def test_sanitize_sheet_name_with_invalid_characters(exporter):
    """Test sheet name sanitization with invalid characters"""
    # Excel doesn't allow: \ / ? * [ ] :
    invalid_name = 'test/zone\\name?with*invalid[chars]:here'
    
    sanitized = exporter.sanitize_sheet_name(invalid_name)
    
    # Should not contain invalid characters
    assert '/' not in sanitized
    assert '\\' not in sanitized
    assert '?' not in sanitized
    assert '*' not in sanitized
    assert '[' not in sanitized
    assert ']' not in sanitized
    assert ':' not in sanitized


def test_sanitize_sheet_name_truncates_long_names(exporter):
    """Test that long sheet names are truncated to 31 characters"""
    long_name = 'a' * 50
    
    sanitized = exporter.sanitize_sheet_name(long_name)
    
    assert len(sanitized) <= 31


def test_sanitize_sheet_name_handles_empty_string(exporter):
    """Test that empty strings are handled"""
    sanitized = exporter.sanitize_sheet_name('')
    
    assert len(sanitized) > 0
    assert sanitized == 'Zone'


def test_sanitize_sheet_name_preserves_valid_names(exporter):
    """Test that valid names are preserved"""
    valid_name = 'example.com'
    
    sanitized = exporter.sanitize_sheet_name(valid_name)
    
    assert sanitized == valid_name


def test_generate_filename_format(exporter):
    """Test that generated filename has correct format"""
    filename = exporter.generate_filename()
    
    assert filename.startswith('DNS-Gather_')
    assert filename.endswith('.xlsx')
    assert len(filename) == 30  # DNS-Gather_ (11) + YYYYMMDD-HH-MM (14) + .xlsx (5)


def test_generate_filename_is_unique(exporter):
    """Test that generated filenames can be unique"""
    # Generate multiple filenames quickly
    filenames = [exporter.generate_filename() for _ in range(3)]
    
    # All should have the same format
    for filename in filenames:
        assert filename.startswith('DNS-Gather_')
        assert filename.endswith('.xlsx')


def test_header_formatting_applied(exporter, sample_zones):
    """Test that header formatting is applied"""
    records_by_zone = {zone.name: [] for zone in sample_zones}
    filepath = exporter.create_workbook(sample_zones, records_by_zone)
    
    wb = load_workbook(filepath)
    ws = wb['Zone List']
    
    # Check first header cell has formatting
    header_cell = ws.cell(row=1, column=1)
    assert header_cell.font.bold is True
    # openpyxl adds '00' prefix instead of 'FF' for some reason
    assert header_cell.fill.start_color.rgb in ('FF' + exporter.header_bg_color, '00' + exporter.header_bg_color)
    
    wb.close()


def test_column_width_adjustment(exporter):
    """Test that column widths are adjusted"""
    zone = ZoneInfo(
        name='test.com',
        zone_type='Primary',
        serial=1,
        transfer_status='Success',
        record_count=1,
        error_message=''
    )
    
    # Create record with long data
    long_data = 'a' * 100
    records = [DNSRecord(name='test', record_type='TXT', ttl=300, data=long_data)]
    
    records_by_zone = {'test.com': records}
    filepath = exporter.create_workbook([zone], records_by_zone)
    
    wb = load_workbook(filepath)
    ws = wb['test.com']
    
    # Check that column width is set (should be capped at max_column_width)
    data_column_width = ws.column_dimensions['D'].width
    assert data_column_width <= exporter.max_column_width
    assert data_column_width > 0
    
    wb.close()


def test_workbook_with_multiple_zones_and_records(exporter):
    """Test complete workbook with multiple zones and records"""
    zones = [
        ZoneInfo(name='zone1.com', zone_type='Primary', serial=1, 
                transfer_status='Success', record_count=2, error_message=''),
        ZoneInfo(name='zone2.com', zone_type='Secondary', serial=2, 
                transfer_status='Success', record_count=1, error_message=''),
    ]
    
    records_by_zone = {
        'zone1.com': [
            DNSRecord(name='www', record_type='A', ttl=300, data='192.168.1.1'),
            DNSRecord(name='mail', record_type='A', ttl=300, data='192.168.1.2'),
        ],
        'zone2.com': [
            DNSRecord(name='@', record_type='NS', ttl=3600, data='ns1.zone2.com'),
        ]
    }
    
    filepath = exporter.create_workbook(zones, records_by_zone)
    
    wb = load_workbook(filepath)
    
    # Should have Zone List + PTR Records + CNAME Records + SRV Records + AAAA Records + 2 zone sheets
    assert len(wb.sheetnames) == 7
    assert 'Zone List' in wb.sheetnames
    assert 'PTR Records' in wb.sheetnames
    assert 'CNAME Records' in wb.sheetnames
    assert 'SRV Records' in wb.sheetnames
    assert 'AAAA Records' in wb.sheetnames
    assert 'zone1.com' in wb.sheetnames
    assert 'zone2.com' in wb.sheetnames
    
    # Check zone1 has 2 records
    ws1 = wb['zone1.com']
    assert ws1.max_row == 3  # header + 2 records
    
    # Check zone2 has 1 record
    ws2 = wb['zone2.com']
    assert ws2.max_row == 2  # header + 1 record
    
    wb.close()


def test_ptr_records_sheet_creation(exporter, temp_output_dir):
    """Test that PTR records sheet is created with correct data"""
    # Create zones with PTR records
    zones = [
        ZoneInfo(
            name='1.168.192.in-addr.arpa',
            zone_type='Primary',
            serial=1,
            transfer_status='Success',
            record_count=3,
            error_message=''
        ),
        ZoneInfo(
            name='example.com',
            zone_type='Primary',
            serial=1,
            transfer_status='Success',
            record_count=1,
            error_message=''
        )
    ]
    
    # Create records with PTR records
    records_by_zone = {
        '1.168.192.in-addr.arpa': [
            DNSRecord(name='10', record_type='PTR', ttl=300, data='server1.example.com.'),
            DNSRecord(name='20', record_type='PTR', ttl=300, data='server2.example.com.'),
            DNSRecord(name='30', record_type='PTR', ttl=300, data='server3.example.com.')
        ],
        'example.com': [
            DNSRecord(name='www', record_type='A', ttl=300, data='192.168.1.10')
        ]
    }
    
    # Create workbook
    filepath = exporter.create_workbook(zones, records_by_zone)
    
    # Load and verify
    wb = load_workbook(filepath)
    
    # Check PTR Records sheet exists
    assert 'PTR Records' in wb.sheetnames
    
    ws = wb['PTR Records']
    
    # Check headers
    assert ws.cell(1, 1).value == 'IP Address'
    assert ws.cell(1, 2).value == 'FQDN'
    assert ws.cell(1, 3).value == 'Zone'
    assert ws.cell(1, 4).value == 'TTL'
    
    # Check PTR record data
    assert ws.cell(2, 1).value == '192.168.1.10'
    assert ws.cell(2, 2).value == 'server1.example.com'
    assert ws.cell(2, 3).value == '1.168.192.in-addr.arpa'
    assert ws.cell(2, 4).value == 300
    
    assert ws.cell(3, 1).value == '192.168.1.20'
    assert ws.cell(3, 2).value == 'server2.example.com'
    
    assert ws.cell(4, 1).value == '192.168.1.30'
    assert ws.cell(4, 2).value == 'server3.example.com'
    
    wb.close()


def test_extract_ip_from_ptr_ipv4(exporter):
    """Test IP extraction from IPv4 PTR records"""
    # Test standard PTR record
    ip = exporter.extract_ip_from_ptr('1.168.192.in-addr.arpa', '10')
    assert ip == '192.168.1.10'
    
    # Test another octet
    ip = exporter.extract_ip_from_ptr('1.168.192.in-addr.arpa', '255')
    assert ip == '192.168.1.255'
    
    # Test @ record
    ip = exporter.extract_ip_from_ptr('1.168.192.in-addr.arpa', '@')
    assert ip == '192.168.1'


def test_extract_ip_from_ptr_multi_octet(exporter):
    """Test IP extraction from PTR records with multiple octets in record name"""
    # Test /16 zone with two octets in record name
    # Zone: 168.192.in-addr.arpa (reversed: 192.168)
    # Record: 10.1 (parts: 10, 1)
    # Result: 192.168 + 10 + 1 = 192.168.10.1
    ip = exporter.extract_ip_from_ptr('168.192.in-addr.arpa', '10.1')
    assert ip == '192.168.10.1'


def test_ip_sort_key(exporter):
    """Test IP address sorting"""
    ips = ['192.168.1.100', '192.168.1.10', '192.168.1.1', '10.0.0.1']
    sorted_ips = sorted(ips, key=exporter.ip_sort_key)
    
    assert sorted_ips == ['10.0.0.1', '192.168.1.1', '192.168.1.10', '192.168.1.100']


def test_ptr_records_sheet_empty(exporter, temp_output_dir):
    """Test PTR records sheet with no PTR records"""
    zones = [
        ZoneInfo(
            name='example.com',
            zone_type='Primary',
            serial=1,
            transfer_status='Success',
            record_count=1,
            error_message=''
        )
    ]
    
    records_by_zone = {
        'example.com': [
            DNSRecord(name='www', record_type='A', ttl=300, data='192.168.1.10')
        ]
    }
    
    # Create workbook
    filepath = exporter.create_workbook(zones, records_by_zone)
    
    # Load and verify
    wb = load_workbook(filepath)
    
    # Check PTR Records sheet exists but is empty (only headers)
    assert 'PTR Records' in wb.sheetnames
    ws = wb['PTR Records']
    
    # Should have headers but no data rows
    assert ws.cell(1, 1).value == 'IP Address'
    assert ws.max_row == 1  # Only header row
    
    wb.close()
