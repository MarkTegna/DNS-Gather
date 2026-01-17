"""Property-based tests for Excel export functionality"""

import pytest
import tempfile
from pathlib import Path
from hypothesis import given, strategies as st, settings
from openpyxl import load_workbook

from dns_gather.excel_exporter import ExcelExporter
from dns_gather.models import ZoneInfo, DNSRecord


@st.composite
def zone_info_strategy(draw):
    """Generate random ZoneInfo objects"""
    name = draw(st.text(min_size=3, max_size=30, alphabet='abcdefghijklmnopqrstuvwxyz0123456789.-').filter(
        lambda x: not x.startswith('.') and not x.endswith('.') and '..' not in x
    ))
    zone_type = draw(st.sampled_from(['Primary', 'Secondary', 'Stub', 'Forward']))
    serial = draw(st.integers(min_value=1, max_value=2147483647))
    record_count = draw(st.integers(min_value=0, max_value=10000))
    transfer_status = draw(st.sampled_from(['Pending', 'Success', 'Failed']))
    error_message = draw(st.text(max_size=100, alphabet='abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789 '))
    
    return ZoneInfo(
        name=name,
        zone_type=zone_type,
        serial=serial,
        transfer_status=transfer_status,
        record_count=record_count,
        error_message=error_message
    )


@st.composite
def dns_record_strategy(draw):
    """Generate random DNSRecord objects"""
    name = draw(st.text(min_size=1, max_size=50, alphabet='abcdefghijklmnopqrstuvwxyz0123456789.-'))
    record_type = draw(st.sampled_from(['A', 'AAAA', 'CNAME', 'MX', 'NS', 'SOA', 'TXT', 'PTR', 'SRV']))
    ttl = draw(st.integers(min_value=0, max_value=86400))
    data = draw(st.text(min_size=1, max_size=100, alphabet='abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789.-_ '))
    
    return DNSRecord(
        name=name,
        record_type=record_type,
        ttl=ttl,
        data=data
    )


# Feature: dns-zone-exporter, Property 8: Excel Workbook Creation
@given(
    zones=st.lists(zone_info_strategy(), min_size=1, max_size=5)
)
@settings(max_examples=10)
def test_workbook_creation_produces_valid_file(zones):
    """
    For any list of zones, workbook creation should produce a valid Excel file
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        exporter = ExcelExporter(output_directory=tmpdir)
        
        # Create empty records dict
        records_by_zone = {zone.name: [] for zone in zones}
        
        # Create workbook
        filepath = exporter.create_workbook(zones, records_by_zone)
        
        # Verify file exists
        assert Path(filepath).exists()
        
        # Verify it's a valid Excel file (close it immediately)
        wb = load_workbook(filepath)
        assert wb is not None
        wb.close()
        
        # Delete file to avoid Windows file locking issues
        Path(filepath).unlink()


# Feature: dns-zone-exporter, Property 9: Zone List Sheet Structure
@given(
    zones=st.lists(zone_info_strategy(), min_size=1, max_size=5)
)
@settings(max_examples=10)
def test_zone_list_sheet_contains_all_zones(zones):
    """
    For any list of zones, the zone list sheet should contain all zones
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        exporter = ExcelExporter(output_directory=tmpdir)
        
        records_by_zone = {zone.name: [] for zone in zones}
        filepath = exporter.create_workbook(zones, records_by_zone)
        
        # Load workbook and check zone list sheet
        wb = load_workbook(filepath)
        ws = wb['Zone List']
        
        # Should have header row + one row per zone
        assert ws.max_row == len(zones) + 1
        
        # Check that all zone names are present
        zone_names_in_sheet = [ws.cell(row=i+2, column=1).value for i in range(len(zones))]
        zone_names_expected = [zone.name for zone in zones]
        
        assert set(zone_names_in_sheet) == set(zone_names_expected)
        
        wb.close()
        Path(filepath).unlink()


# Feature: dns-zone-exporter, Property 9: Zone List Sheet Structure
@given(
    zones=st.lists(zone_info_strategy(), min_size=1, max_size=3)
)
@settings(max_examples=10)
def test_zone_list_sheet_has_correct_headers(zones):
    """
    For any list of zones, the zone list sheet should have correct headers
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        exporter = ExcelExporter(output_directory=tmpdir)
        
        records_by_zone = {zone.name: [] for zone in zones}
        filepath = exporter.create_workbook(zones, records_by_zone)
        
        # Load workbook and check headers
        wb = load_workbook(filepath)
        ws = wb['Zone List']
        
        # Check headers
        expected_headers = ['Zone Name', 'Type', 'Serial', 'Record Count', 'Transfer Status', 'Error Message']
        actual_headers = [ws.cell(row=1, column=i+1).value for i in range(6)]
        
        assert actual_headers == expected_headers
        
        wb.close()
        Path(filepath).unlink()


# Feature: dns-zone-exporter, Property 10: Zone Sheet Creation
@given(
    zones=st.lists(zone_info_strategy(), min_size=1, max_size=3)
)
@settings(max_examples=10)
def test_workbook_contains_sheet_for_each_zone(zones):
    """
    For any list of zones, the workbook should contain a sheet for each zone
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        exporter = ExcelExporter(output_directory=tmpdir)
        
        records_by_zone = {zone.name: [] for zone in zones}
        filepath = exporter.create_workbook(zones, records_by_zone)
        
        # Load workbook
        wb = load_workbook(filepath)
        
        # Should have Zone List sheet + one sheet per zone
        assert len(wb.sheetnames) == len(zones) + 1
        assert 'Zone List' in wb.sheetnames
        
        wb.close()
        Path(filepath).unlink()


# Feature: dns-zone-exporter, Property 11: Zone Sheet Structure
@given(
    zone=zone_info_strategy(),
    records=st.lists(dns_record_strategy(), min_size=1, max_size=10)
)
@settings(max_examples=10)
def test_zone_sheet_contains_all_records(zone, records):
    """
    For any zone with records, the zone sheet should contain all records
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        exporter = ExcelExporter(output_directory=tmpdir)
        
        records_by_zone = {zone.name: records}
        filepath = exporter.create_workbook([zone], records_by_zone)
        
        # Load workbook and find zone sheet
        wb = load_workbook(filepath)
        
        # Get zone sheet (sanitized name)
        sanitized_name = exporter.sanitize_sheet_name(zone.name)
        ws = wb[sanitized_name]
        
        # Should have header row + one row per record
        assert ws.max_row == len(records) + 1
        
        wb.close()
        Path(filepath).unlink()


# Feature: dns-zone-exporter, Property 11: Zone Sheet Structure
@given(
    zone=zone_info_strategy(),
    records=st.lists(dns_record_strategy(), min_size=0, max_size=5)
)
@settings(max_examples=10)
def test_zone_sheet_has_correct_headers(zone, records):
    """
    For any zone, the zone sheet should have correct headers
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        exporter = ExcelExporter(output_directory=tmpdir)
        
        records_by_zone = {zone.name: records}
        filepath = exporter.create_workbook([zone], records_by_zone)
        
        # Load workbook
        wb = load_workbook(filepath)
        sanitized_name = exporter.sanitize_sheet_name(zone.name)
        ws = wb[sanitized_name]
        
        # Check headers
        expected_headers = ['Name', 'Type', 'TTL', 'Data']
        actual_headers = [ws.cell(row=1, column=i+1).value for i in range(4)]
        
        assert actual_headers == expected_headers
        
        wb.close()
        Path(filepath).unlink()


# Feature: dns-zone-exporter, Property 12: Sheet Name Sanitization
@given(
    name=st.text(min_size=1, max_size=100, alphabet='abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789.-_/\\*?:[]')
)
@settings(max_examples=100)
def test_sanitized_sheet_names_are_valid(name):
    """
    For any string, sanitized sheet names should be valid for Excel
    """
    exporter = ExcelExporter()
    
    sanitized = exporter.sanitize_sheet_name(name)
    
    # Should not be empty
    assert len(sanitized) > 0
    
    # Should be <= 31 characters
    assert len(sanitized) <= 31
    
    # Should not contain invalid characters
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    for char in invalid_chars:
        assert char not in sanitized


# Feature: dns-zone-exporter, Property 12: Sheet Name Sanitization
@given(
    name=st.text(min_size=1, max_size=50, alphabet='abcdefghijklmnopqrstuvwxyz0123456789.-')
)
@settings(max_examples=100)
def test_sanitized_names_preserve_valid_characters(name):
    """
    For any string with valid characters, sanitization should preserve them
    """
    exporter = ExcelExporter()
    
    sanitized = exporter.sanitize_sheet_name(name)
    
    # Should preserve alphanumeric, dots, and hyphens
    for char in sanitized:
        assert char.isalnum() or char in '.-_ '


# Feature: dns-zone-exporter, Property 13: Timestamped Filename Format
def test_filename_format_is_correct():
    """
    For any generated filename, it should follow the correct format
    """
    exporter = ExcelExporter()
    
    for _ in range(10):
        filename = exporter.generate_filename()
        
        # Should start with DNS-Gather_
        assert filename.startswith('DNS-Gather_')
        
        # Should end with .xlsx
        assert filename.endswith('.xlsx')
        
        # Should contain timestamp in YYYYMMDD-HH-MM format
        # Extract timestamp part
        timestamp_part = filename.replace('DNS-Gather_', '').replace('.xlsx', '')
        
        # Should be 14 characters (YYYYMMDD-HH-MM)
        assert len(timestamp_part) == 14
        
        # Should have hyphens at correct positions
        assert timestamp_part[8] == '-'
        assert timestamp_part[11] == '-'


# Feature: dns-zone-exporter, Property 13: Timestamped Filename Format
def test_filename_contains_valid_timestamp():
    """
    For any generated filename, the timestamp should be valid
    """
    exporter = ExcelExporter()
    
    filename = exporter.generate_filename()
    
    # Extract timestamp
    timestamp_part = filename.replace('DNS-Gather_', '').replace('.xlsx', '')
    
    # Parse components
    year = int(timestamp_part[0:4])
    month = int(timestamp_part[4:6])
    day = int(timestamp_part[6:8])
    hour = int(timestamp_part[9:11])
    minute = int(timestamp_part[12:14])
    
    # Validate ranges
    assert 2020 <= year <= 2100
    assert 1 <= month <= 12
    assert 1 <= day <= 31
    assert 0 <= hour <= 23
    assert 0 <= minute <= 59
