"""Excel export functionality for DNS-Gather"""

from datetime import datetime
from pathlib import Path
from typing import List
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from dns_gather.models import ZoneInfo, DNSRecord


class ExcelExporter:
    """Exports DNS zone data to Excel workbooks"""
    
    def __init__(self, output_directory: str = './Reports'):
        """
        Initialize Excel Exporter
        
        Args:
            output_directory: Directory for output files
        """
        self.output_directory = Path(output_directory)
        self.output_directory.mkdir(parents=True, exist_ok=True)
        
        # Excel formatting settings
        self.header_bg_color = '4472C4'
        self.header_font_color = 'FFFFFF'
        self.max_column_width = 50
    
    def create_workbook(self, zones: List[ZoneInfo], records_by_zone: dict) -> str:
        """
        Create Excel workbook with zone data
        
        Args:
            zones: List of ZoneInfo objects
            records_by_zone: Dictionary mapping zone names to lists of DNSRecord objects
        
        Returns:
            Path to created workbook file
        """
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create zone list sheet
        self.create_zone_list_sheet(wb, zones)
        
        # Create consolidation sheets for specific record types
        self.create_ptr_records_sheet(wb, zones, records_by_zone)
        self.create_cname_records_sheet(wb, zones, records_by_zone)
        self.create_srv_records_sheet(wb, zones, records_by_zone)
        self.create_aaaa_records_sheet(wb, zones, records_by_zone)
        
        # Create individual zone sheets
        for zone in zones:
            records = records_by_zone.get(zone.name, [])
            self.create_zone_sheet(wb, zone, records)
        
        # Generate filename with timestamp
        filename = self.generate_filename()
        filepath = self.output_directory / filename
        
        # Save workbook
        wb.save(filepath)
        
        return str(filepath)
    
    def create_zone_list_sheet(self, wb: Workbook, zones: List[ZoneInfo]) -> None:
        """
        Create zone list summary sheet
        
        Args:
            wb: Workbook object
            zones: List of ZoneInfo objects
        """
        ws = wb.create_sheet('Zone List', 0)
        
        # Headers
        headers = ['Zone Name', 'Type', 'Serial', 'Record Count', 'Transfer Status', 'Error Message']
        ws.append(headers)
        
        # Apply header formatting
        self.apply_formatting(ws, len(headers))
        
        # Add zone data
        for zone in zones:
            ws.append([
                zone.name,
                zone.zone_type,
                zone.serial,
                zone.record_count,
                zone.transfer_status,
                zone.error_message
            ])
        
        # Auto-adjust column widths
        self.auto_adjust_columns(ws)
    
    def create_ptr_records_sheet(self, wb: Workbook, zones: List[ZoneInfo], records_by_zone: dict) -> None:
        """
        Create consolidated PTR records sheet with IP addresses and FQDNs
        
        Args:
            wb: Workbook object
            zones: List of ZoneInfo objects
            records_by_zone: Dictionary mapping zone names to lists of DNSRecord objects
        """
        ws = wb.create_sheet('PTR Records', 1)
        
        # Headers
        headers = ['IP Address', 'FQDN', 'Zone', 'TTL']
        ws.append(headers)
        
        # Apply header formatting
        self.apply_formatting(ws, len(headers))
        
        # Collect all PTR records
        ptr_records = []
        
        for zone in zones:
            records = records_by_zone.get(zone.name, [])
            
            # Check if this is a reverse zone (in-addr.arpa or ip6.arpa)
            is_reverse_zone = zone.name.endswith('.in-addr.arpa') or zone.name.endswith('.ip6.arpa')
            
            for record in records:
                if record.record_type == 'PTR':
                    # Extract IP address from zone name and record name
                    ip_address = self.extract_ip_from_ptr(zone.name, record.name)
                    
                    # FQDN is the PTR data (remove trailing dot)
                    fqdn = record.data.rstrip('.')
                    
                    ptr_records.append({
                        'ip': ip_address,
                        'fqdn': fqdn,
                        'zone': zone.name,
                        'ttl': record.ttl
                    })
        
        # Sort by IP address
        ptr_records.sort(key=lambda x: self.ip_sort_key(x['ip']))
        
        # Add PTR record data
        for ptr in ptr_records:
            ws.append([
                ptr['ip'],
                ptr['fqdn'],
                ptr['zone'],
                ptr['ttl']
            ])
        
        # Auto-adjust column widths
        self.auto_adjust_columns(ws)
    
    def create_cname_records_sheet(self, wb: Workbook, zones: List[ZoneInfo], records_by_zone: dict) -> None:
        """
        Create consolidated CNAME records sheet
        
        Args:
            wb: Workbook object
            zones: List of ZoneInfo objects
            records_by_zone: Dictionary mapping zone names to lists of DNSRecord objects
        """
        ws = wb.create_sheet('CNAME Records')
        
        # Headers
        headers = ['Name', 'Target', 'Zone', 'TTL']
        ws.append(headers)
        
        # Apply header formatting
        self.apply_formatting(ws, len(headers))
        
        # Collect all CNAME records
        cname_records = []
        
        for zone in zones:
            records = records_by_zone.get(zone.name, [])
            
            for record in records:
                if record.record_type == 'CNAME':
                    # Build FQDN for the name
                    if record.name == '@':
                        fqdn = zone.name
                    else:
                        fqdn = f"{record.name}.{zone.name}"
                    
                    # Target is the CNAME data (remove trailing dot)
                    target = record.data.rstrip('.')
                    
                    cname_records.append({
                        'name': fqdn,
                        'target': target,
                        'zone': zone.name,
                        'ttl': record.ttl
                    })
        
        # Sort by name
        cname_records.sort(key=lambda x: x['name'].lower())
        
        # Add CNAME record data
        for cname in cname_records:
            ws.append([
                cname['name'],
                cname['target'],
                cname['zone'],
                cname['ttl']
            ])
        
        # Auto-adjust column widths
        self.auto_adjust_columns(ws)
    
    def create_srv_records_sheet(self, wb: Workbook, zones: List[ZoneInfo], records_by_zone: dict) -> None:
        """
        Create consolidated SRV records sheet
        
        Args:
            wb: Workbook object
            zones: List of ZoneInfo objects
            records_by_zone: Dictionary mapping zone names to lists of DNSRecord objects
        """
        ws = wb.create_sheet('SRV Records')
        
        # Headers
        headers = ['Service', 'Priority', 'Weight', 'Port', 'Target', 'Zone', 'TTL']
        ws.append(headers)
        
        # Apply header formatting
        self.apply_formatting(ws, len(headers))
        
        # Collect all SRV records
        srv_records = []
        
        for zone in zones:
            records = records_by_zone.get(zone.name, [])
            
            for record in records:
                if record.record_type == 'SRV':
                    # Build FQDN for the service name
                    if record.name == '@':
                        service = zone.name
                    else:
                        service = f"{record.name}.{zone.name}"
                    
                    # Parse SRV data: "priority weight port target"
                    # Example: "10 60 5060 sipserver.example.com."
                    parts = record.data.split(None, 3)
                    
                    if len(parts) >= 4:
                        priority = parts[0]
                        weight = parts[1]
                        port = parts[2]
                        target = parts[3].rstrip('.')
                    else:
                        # Malformed SRV record, use raw data
                        priority = weight = port = ''
                        target = record.data.rstrip('.')
                    
                    srv_records.append({
                        'service': service,
                        'priority': priority,
                        'weight': weight,
                        'port': port,
                        'target': target,
                        'zone': zone.name,
                        'ttl': record.ttl
                    })
        
        # Sort by service name
        srv_records.sort(key=lambda x: x['service'].lower())
        
        # Add SRV record data
        for srv in srv_records:
            ws.append([
                srv['service'],
                srv['priority'],
                srv['weight'],
                srv['port'],
                srv['target'],
                srv['zone'],
                srv['ttl']
            ])
        
        # Auto-adjust column widths
        self.auto_adjust_columns(ws)
    
    def create_aaaa_records_sheet(self, wb: Workbook, zones: List[ZoneInfo], records_by_zone: dict) -> None:
        """
        Create consolidated AAAA (IPv6) records sheet
        
        Args:
            wb: Workbook object
            zones: List of ZoneInfo objects
            records_by_zone: Dictionary mapping zone names to lists of DNSRecord objects
        """
        ws = wb.create_sheet('AAAA Records')
        
        # Headers
        headers = ['Name', 'IPv6 Address', 'Zone', 'TTL']
        ws.append(headers)
        
        # Apply header formatting
        self.apply_formatting(ws, len(headers))
        
        # Collect all AAAA records
        aaaa_records = []
        
        for zone in zones:
            records = records_by_zone.get(zone.name, [])
            
            for record in records:
                if record.record_type == 'AAAA':
                    # Build FQDN for the name
                    if record.name == '@':
                        fqdn = zone.name
                    else:
                        fqdn = f"{record.name}.{zone.name}"
                    
                    # IPv6 address is the data
                    ipv6_address = record.data
                    
                    aaaa_records.append({
                        'name': fqdn,
                        'ipv6': ipv6_address,
                        'zone': zone.name,
                        'ttl': record.ttl
                    })
        
        # Sort by name
        aaaa_records.sort(key=lambda x: x['name'].lower())
        
        # Add AAAA record data
        for aaaa in aaaa_records:
            ws.append([
                aaaa['name'],
                aaaa['ipv6'],
                aaaa['zone'],
                aaaa['ttl']
            ])
        
        # Auto-adjust column widths
        self.auto_adjust_columns(ws)
    
    def extract_ip_from_ptr(self, zone_name: str, record_name: str) -> str:
        """
        Extract IP address from PTR zone name and record name
        
        Args:
            zone_name: Reverse zone name (e.g., "1.168.192.in-addr.arpa")
            record_name: PTR record name (e.g., "10" or "@")
        
        Returns:
            IP address string (e.g., "192.168.1.10")
        """
        try:
            if zone_name.endswith('.in-addr.arpa'):
                # IPv4 reverse zone
                # Remove .in-addr.arpa suffix
                zone_parts = zone_name.replace('.in-addr.arpa', '').split('.')
                
                # Reverse the octets (they're in reverse order in DNS)
                zone_parts.reverse()
                
                # Handle record name
                if record_name == '@':
                    # @ means the zone itself - this shouldn't typically have a PTR
                    # but if it does, use the zone as-is
                    return '.'.join(zone_parts)
                else:
                    # Record name contains the remaining octets
                    record_parts = record_name.split('.')
                    # Combine zone parts with record parts
                    all_parts = zone_parts + record_parts
                    return '.'.join(all_parts)
            
            elif zone_name.endswith('.ip6.arpa'):
                # IPv6 reverse zone
                # This is more complex - for now, return the raw format
                # Full IPv6 PTR reconstruction would require more logic
                zone_parts = zone_name.replace('.ip6.arpa', '').split('.')
                zone_parts.reverse()
                
                if record_name == '@':
                    ipv6_hex = ''.join(zone_parts)
                else:
                    record_parts = record_name.split('.')
                    record_parts.reverse()
                    ipv6_hex = ''.join(record_parts + zone_parts)
                
                # Format as IPv6 (insert colons every 4 characters)
                ipv6_formatted = ':'.join([ipv6_hex[i:i+4] for i in range(0, len(ipv6_hex), 4)])
                return ipv6_formatted
            
            else:
                # Not a reverse zone - return as-is
                return f"{record_name}.{zone_name}"
        
        except Exception:
            # If parsing fails, return a combined string
            return f"{record_name}.{zone_name}"
    
    def ip_sort_key(self, ip_str: str) -> tuple:
        """
        Generate sort key for IP address (handles both IPv4 and IPv6)
        
        Args:
            ip_str: IP address string
        
        Returns:
            Tuple for sorting
        """
        try:
            # Try to parse as IPv4
            if '.' in ip_str and ':' not in ip_str:
                parts = ip_str.split('.')
                return tuple(int(p) if p.isdigit() else 0 for p in parts[:4])
            else:
                # IPv6 or other - sort alphabetically
                return (ip_str,)
        except Exception:
            return (ip_str,)
    
    def create_zone_sheet(self, wb: Workbook, zone: ZoneInfo, records: List[DNSRecord]) -> None:
        """
        Create sheet for individual zone
        
        Args:
            wb: Workbook object
            zone: ZoneInfo object
            records: List of DNSRecord objects for this zone
        """
        # Sanitize sheet name for Excel compatibility
        sheet_name = self.sanitize_sheet_name(zone.name)
        
        ws = wb.create_sheet(sheet_name)
        
        # Headers
        headers = ['Name', 'Type', 'TTL', 'Data']
        ws.append(headers)
        
        # Apply header formatting
        self.apply_formatting(ws, len(headers))
        
        # Add record data
        for record in records:
            ws.append([
                record.name,
                record.record_type,
                record.ttl,
                record.data
            ])
        
        # Auto-adjust column widths
        self.auto_adjust_columns(ws)
    
    def sanitize_sheet_name(self, name: str) -> str:
        r"""
        Sanitize zone name for Excel sheet name compatibility
        
        Excel sheet names cannot contain: \ / ? * [ ] :
        Excel sheet names must be <= 31 characters
        
        Args:
            name: Original zone name
        
        Returns:
            Sanitized sheet name
        """
        # Replace invalid characters with underscore
        sanitized = re.sub(r'[\\/*?:\[\]]', '_', name)
        
        # Truncate to 31 characters
        if len(sanitized) > 31:
            sanitized = sanitized[:31]
        
        # Remove trailing/leading spaces and dots
        sanitized = sanitized.strip('. ')
        
        # Ensure not empty
        if not sanitized:
            sanitized = 'Zone'
        
        return sanitized
    
    def apply_formatting(self, ws, num_columns: int) -> None:
        """
        Apply formatting to header row
        
        Args:
            ws: Worksheet object
            num_columns: Number of columns to format
        """
        # Header row styling
        header_fill = PatternFill(start_color=self.header_bg_color, 
                                  end_color=self.header_bg_color, 
                                  fill_type='solid')
        header_font = Font(color=self.header_font_color, bold=True)
        header_alignment = Alignment(horizontal='left', vertical='center')
        
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
    
    def auto_adjust_columns(self, ws) -> None:
        """
        Auto-adjust column widths based on content
        
        Args:
            ws: Worksheet object
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set column width with max limit
            adjusted_width = min(max_length + 2, self.max_column_width)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_filename(self) -> str:
        """
        Generate timestamped filename
        
        Returns:
            Filename in format: DNS-Gather_YYYYMMDD-HH-MM.xlsx
        """
        timestamp = datetime.now().strftime('%Y%m%d-%H-%M')
        return f'DNS-Gather_{timestamp}.xlsx'
